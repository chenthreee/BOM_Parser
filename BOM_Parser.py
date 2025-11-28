from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox
from footprint.footprint import initialize_set, add_to_set, remove_from_set, get_footprint_set, any_footprint_in_string
from preProcess.KB_preProcess import KB_manufacturer_process
from preProcess.MC_preProcess import MC_preProcess
from preProcess.mergePreprocess import merge_preprocess
from query import query_online, token_access, refresh_excel
from test import Ui_Dialog
import pandas as pd
from db_model import K3Data
from typing import Dict
import openpyxl
from openpyxl.styles import PatternFill
import re

# 初始化封装集合
initialize_set(
    {'DIP', 'DIL', 'SIP', 'SIL', 'ZIP', 'QFP', 'SOJ', 'PLCC', 'CLCC', 'TSOC', 'PGA', 'BGA', 'SMD', 'SOP', 'SOIC',
     'CERPAC', 'DMP', 'SC70', 'SOT', 'TO', 'QFN', '0201', '0402', '0603', '0805', '1206', '1210', '1812', '2010',
     '2512'})


def remove_last_word(s):
    # 使用正则表达式处理中英文空格，从右边分割一次
    parts = re.split(r'[ 　]+', s)
    return ' '.join(parts[:-1]) if len(parts) > 1 else ''


def text_to_val(lineEdit_text) -> int:
    try:
        assert len(lineEdit_text) > 0
        return int(lineEdit_text)
    except AssertionError:
        print("AssertionError : {lineEdit_text}")
    except ValueError:
        print("cannot convert text to number : {lineEdit_text}")


# 通过仅用料号强匹配进行查询
def query_part_info_weak_match(part_code, customer):
    """
    使用仅用料号强匹配在K3数据库中进行查询（不匹配客户）
    :param part_code:料号
    :param customer:客户编码（保留参数但不使用）
    :return:该料号在K3中的规格描述列表
    """
    # 去除料号首尾空格
    part_code = part_code.strip()

    all_results = K3Data.select()
    query_results = []
    for r in all_results:
        parts = re.split(r'[ 　]+', r.specification)
        if len(parts) >= 2 and parts[1] == part_code:
            query_results.append(r)

    if len(query_results) == 0:
        return []
    else:
        results_list = []
        for query_result in query_results:
            results_list.append({
                'k3code': query_result.k3code,
                'type_name': query_result.type_name,
                'specification': query_result.specification,
                'match_type': '仅用料号强匹配'  # 标记为仅用料号强匹配
            })
        return results_list


# 通过料号进行查询，返回K3中该料号的相关信息
def query_part_info_by_partcode(part_code):
    """
    使用料号在K3数据库中进行查询，返回对应的规格描述
    :param part_code:料号
    :return:该客户、该料号在K3中的规格描述列表
    """
    # 去除料号首尾空格
    part_code = part_code.strip()

    query_results = K3Data.select().where(K3Data.specification.contains(part_code))

    if len(query_results) == 0:  # case 1：如果没有查询结果
        return []
    else:  # 返回所有查询结果的列表
        results_list = []
        for query_result in query_results:
            results_list.append({
                'k3code': None,  # 客户不匹配，不返回K3 编码
                'type_name': query_result.type_name,
                'specification': query_result.specification,
                'match_type': '仅用料号弱匹配'
            })
        return results_list


def query_part_info(part_code, customer):
    """
    使用料号、客户编码在K3数据库中进行查询，返回对应的规格描述
    :param part_code:料号
    :param customer:客户编码
    :return:该客户、该料号在K3中的规格描述列表
    """
    # 去除料号首尾空格
    part_code = part_code.strip()

    # 根据 料号以及客户编码进行查询
    print(f"开始数据库查询: part_code={part_code}, customer={customer}")

    #改成这样 客户名称使用强匹配
    # query_results = K3Data.select().where(K3Data.specification.contains(part_code) &
    #                                       K3Data.specification.contains(customer))

    #强查找了
    all_results = K3Data.select().where(K3Data.specification.startswith(customer))
    query_results = []
    for r in all_results:
        parts = re.split(r'[ 　]+', r.specification)
        if len(parts) >= 2 and parts[1] == part_code:
            query_results.append(r)

    result_count = len(query_results)
    print(f"查询到 {result_count} 条结果")

    if len(query_results) >= 1:  # 如果有查询结果，返回所有结果的列表
        results_list = []
        for query_result in query_results:
            results_list.append({
                'k3code': query_result.k3code,
                'type_name': query_result.type_name,
                'specification': query_result.specification
            })
        return results_list
    elif len(query_results) == 0:  # 如果没有查询结果，先尝试仅用料号强匹配
        print(f"无查询结果，尝试仅用料号强匹配")
        weak_results = query_part_info_weak_match(part_code, customer)
        if len(weak_results) > 0:
            return weak_results
        # 料号强匹配也没找到，最后尝试只用料号弱匹配来查
        print(f"料号强匹配无结果，放宽条件只用料号弱匹配查询")
        return query_part_info_by_partcode(part_code)
    # print('{} {} on {}'.format(note.id, note.text, note.created))


class MyMainWindow(QMainWindow, Ui_Dialog):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        # 原始BOM的元数据信息
        self.customer_code = ""
        self.initial_row_line = -1
        self.end_row_line = -1
        self.part_code_col = -1
        self.manufacturer_col = -1
        self.designator_col = -1
        self.quantity_col = -1
        self.description_col = -1
        self.footprint_col = -1
        self.alternative1_code_col = -1
        self.alternative2_code_col = -1
        self.alternative1_manufacturer_col = -1
        self.alternative2_manufacturer_col = -1

    def on_click(self):
        self.customer_code = self.lineEdit_customer_code.text()
        if self.lineEdit_initial_row.text() == '':
            self.messageBox_info('起始行为空')
        else:
            self.initial_row_line = text_to_val(self.lineEdit_initial_row.text())

        if self.lineEdit_end_row.text() == '':
            self.messageBox_info('结束行为空')
        else:
            self.end_row_line = text_to_val(self.lineEdit_end_row.text())

        if self.lineEdit_part_code.text() == '':
            self.messageBox_info('料号列为空')
        else:
            self.part_code_col = ord((self.lineEdit_part_code.text())[0]) - ord('A')

        if self.lineEdit_designator.text() == '':
            self.messageBox_info('位号列为空')
        else:
            self.designator_col = ord((self.lineEdit_designator.text())[0]) - ord('A')

        if 'merge' not in self.customer_code:
            if self.lineEdit_quantity.text() == '':
                self.messageBox_info('用量列为空')
            else:
                self.quantity_col = ord((self.lineEdit_quantity.text())[0]) - ord('A')

        if self.lineEdit_description.text() != '':
            self.description_col = ord((self.lineEdit_description.text())[0]) - ord('A')
        if self.lineEdit_manufacturer.text() != '':
            self.manufacturer_col = ord((self.lineEdit_manufacturer.text())[0]) - ord('A')  # 制造商列
        if self.lineEdit_alternative_part_code_1.text() != '':
            self.alternative1_code_col = ord((self.lineEdit_alternative_part_code_1.text())[0]) - ord('A')  # 替代料号1列号
            self.alternative2_code_col = ord((self.lineEdit_alternative_part_code_2.text())[0]) - ord('A')  # 替代料号2列号
            self.alternative1_manufacturer_col = ord((self.lineEdit_alternative_manufacturer_1.text())[0]) - ord(
                'A')  # 替代料号1厂商列号
            self.alternative2_manufacturer_col = ord((self.lineEdit_alternative_manufacturer_2.text())[0]) - ord(
                'A')  # 替代料号2厂商列号

        # 选填字段处理,可填可不填
        if self.lineEdit_footprint.text() != '':
            self.footprint_col = ord((self.lineEdit_footprint.text())[0]) - ord('A')  # 封装列

        self.DB_process()
        self.raw_BOM_copy_process()
        self.raw_BOM_parse_process()

        # 只有当替代料号列被正确配置时(不为-1),才执行替代料号处理
        if self.alternative1_code_col != -1 and self.alternative1_code_col != self.part_code_col:
            # 先复制target.xlsx到output.xlsx
            import shutil
            shutil.copy('target.xlsx', 'output.xlsx')
            refresh_excel('output.xlsx')
            self.alternative_process()

        # 重命名输出文件
        self.rename_output_file()

        #self.online_query()  # 在线查询

    def rename_output_file(self):
        """
        将生成的工程BOM文件重命名为 原文件名+自动生成工程BOM.xlsx
        并添加颜色标记
        """
        import os
        import shutil

        # 确定要重命名的源文件
        if self.alternative1_code_col != -1 and self.alternative1_code_col != self.part_code_col:
            source_file = 'output.xlsx'
        else:
            source_file = 'target.xlsx'

        # 获取原始BOM文件路径
        bom_file_path = self.BOMInput_textEdit.input_path_file

        # 生成新文件名
        prefix = '自动生成工程BOM'
        name_part, extension = bom_file_path.rsplit('.', 1)
        if extension == 'xls':
            extension = 'xlsx'
        new_filename = name_part + prefix + '.' + extension

        # 复制并重命名文件
        try:
            shutil.copy(source_file, new_filename)
            print(f"工程BOM已保存为: {new_filename}")

            # 添加颜色标记
            self.color_mark(new_filename)

            print("process complete")
        except Exception as e:
            print(f"重命名文件时出错: {e}")

    def color_mark(self, file_path):
        """
        为Excel文件添加颜色标记
        规则:
        - 黄绿色(#99CC00): 当第3列(C列,元件类型)为空时
        - 淡灰色(#D3D3D3): O列标记为'仅用料号弱匹配'
        - 浅蓝色(#ADD8E6): 当该行是多条查询结果中的一条时(O列标记为'multi'或'multi-xxx')
        - 橙色(#FFA500): O列标记为'仅用料号强匹配'
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # 定义颜色填充
            light_green_fill = PatternFill(start_color="99CC00", end_color="99CC00", fill_type="solid")  # 黄绿色
            grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 淡灰色
            light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # 浅蓝色
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 橙色

            # 从第5行开始遍历(跳过前4行表头)
            for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                cell_b = row[1].value  # 第2列(B列,K3 code)
                cell_c = row[2].value  # 第3列(C列,元件类型)

                # 检查第15列(索引14)是否存在且标记为'multi'或包含'multi'
                cell_o = None
                if len(row) > 14:
                    cell_o = row[14].value  # 第15列(O列,多结果标记/匹配类型标记)

                # 检查是否为multi行(包括'multi'和'multi-xxx'格式)
                is_multi = cell_o and isinstance(cell_o, str) and cell_o.startswith('multi')

                if is_multi:  # 优先标记多结果行
                    for cell in row:
                        cell.fill = light_blue_fill  # 标记为浅蓝色
                elif cell_o == '仅用料号弱匹配':  # 仅用料号弱匹配
                    for cell in row:
                        cell.fill = grey_fill  # 标记为灰色
                elif cell_o == '仅用料号强匹配':  # 仅用料号强匹配
                    for cell in row:
                        cell.fill = orange_fill  # 标记为橙色
                elif cell_c is None:  # 如果元件类型为空
                    for cell in row:
                        cell.fill = light_green_fill  # 标记为黄绿色
                elif cell_b is None and cell_c is not None:  # 如果类型有值但K3 code为空(已经被上面的条件覆盖了，这个可能不会执行到)
                    for cell in row:
                        cell.fill = grey_fill  # 标记为淡灰色

            # 保存修改后的Excel文件
            wb.save(file_path)
            print(f"颜色标记已添加")
        except Exception as e:
            print(f"添加颜色标记时出错: {e}")

    def messageBox_info(self, info):
        msg_box = QMessageBox()

        msg_box.setText(info)
        msg_box.setWindowTitle("提示")
        msg_box.setStandardButtons(QMessageBox.Ok)
        retval = msg_box.exec()

    def raw_BOM_copy_process(self):
        """
        对原始BOM中不需要解析的字段直接复制粘贴
        :return:
        """
        if self.customer_code == 'MC':  # MC 预处理
            self.BOMInput_textEdit.input_path_file, me_content_len = MC_preProcess(
                self.BOMInput_textEdit.input_path_file)
            self.end_row_line = self.initial_row_line + me_content_len
        if 'merge' in self.customer_code:  # 坐标BOM预处理
            self.BOMInput_textEdit.input_path_file, self.quantity_col = merge_preprocess(
                self.BOMInput_textEdit.input_path_file, self.initial_row_line - 1, self.part_code_col)
        print(self.BOMInput_textEdit.input_path_file)
        excel_file_path = self.BOMInput_textEdit.input_path_file
        df = pd.read_excel(excel_file_path)
        start_row = self.initial_row_line - 2  # 指定起始行索引（从0开始）
        end_row = self.end_row_line - 1
        designator_list = df.iloc[start_row:end_row, self.designator_col].tolist()  # 位号获取
        if self.footprint_col != -1:
            footprint_list = df.iloc[start_row:end_row, self.footprint_col].tolist()  # 封装获取
        print(self.quantity_col)
        quantity_list = df.iloc[start_row:end_row, self.quantity_col].tolist()  # 用量获取

        description_list = df.iloc[start_row:end_row, self.description_col].tolist()  # 描述获取
        manufacturer_list = []
        if self.manufacturer_col != -1:
            manufacturer_list = df.iloc[start_row:end_row, self.manufacturer_col].tolist()  # 制造商获取

        if self.customer_code == 'KB':  # KB预处理
            manufacturer_list = KB_manufacturer_process(description_list)

        refresh_excel('target.xlsx')
        workbook = openpyxl.load_workbook('target.xlsx')

        sheet = workbook.active

        start_row = 5  # 开始写的行

        for i, value in enumerate(designator_list):
            # 位号转换成BQC格式
            if ',' in str(value):
                sheet['H' + str(start_row + i)] = str(value).replace(' ', '').replace(',', ' ')
            else:
                sheet['H' + str(start_row + i)] = str(value)

        # for i, value in enumerate(footprint_list):   #封装暂时不用写
        #     sheet['G' + str(start_row + i)] = value
        for i, value in enumerate(quantity_list):
            sheet['G' + str(start_row + i)] = value

        if self.manufacturer_col != -1:  # 制造商如果原始BOM 没有单独的列，那么可以不写
            for i, value in enumerate(manufacturer_list):
                sheet['I' + str(start_row + i)] = value
        if self.customer_code == 'KB':
            for i, value in enumerate(manufacturer_list):
                sheet['I' + str(start_row + i)] = value
        for i in range(len(designator_list)):
            sheet['A' + str(start_row + i)] = i + 1

        # 保存Excel文件
        workbook.save('target.xlsx')

    # 强查找
    # I chose HKU because its MSc in AI combines strong theoretical training with practical, application-oriented learning.
    # The programme is jointly offered by the Departments of Mathematics and Computer Science, which gives me a solid 
    # foundation in machine learning as well as the opportunity to work on real-world AI projects through the capstone project. 
    # In addition, HKU’s highly international environment will allow me to learn from experts and peers from diverse backgrounds. 
    # Overall, HKU aligns perfectly with my academic interests and long-term career goals. 

    # AI still faces several key challenges.
    # First, many deep learning models are still black boxes, and we often cannot clearly explain how they make decisions. 
    # This becomes a major concern in high-stakes fields such as healthcare and finance, where transparency is essential.
    # Second, deploying AI systems in real-world environments is difficult because models must be stable, reliable, and efficient
    # in terms of computation, especially on resource-limited devices.
    # Third, privacy protection is also a major challenge, as AI systems often rely on large amounts of user data, which requires
    # responsible and secure handling. 

    # xlsx格式原始BOM处理
    def raw_BOM_parse_process(self):
        """
        对原始BOM中需要解析的字段进行解析处理
        :return:
        """
        print(self.BOMInput_textEdit.input_path_file)
        excel_file_path = self.BOMInput_textEdit.input_path_file
        df = pd.read_excel(excel_file_path)
        column_index = self.part_code_col
        start_row = self.initial_row_line - 2  # 指定起始行索引（从0开始）
        end_row = self.end_row_line - 1
        # column_data = df.iloc[start_row:, column_index].values
        part_code_column = df.iloc[start_row:end_row, column_index].tolist()  # 获取需要查询的料号

        print("查询数量", len(part_code_column))

        ########如果料号和描述混杂成一列，需要通过解析另加功能处理 On pending ############

        # 用于记录每个料号的查询结果列表
        multi_results_map = {}  # {part_index: [result1, result2, ...]}

        for i in range(len(part_code_column)):
            if pd.isnull(part_code_column[i]):  # 防止料号为空
                part_code_column[i] = "料号为空"

            if isinstance(part_code_column[i], (int, float)):  # 将数字类型转换为字符串类型，便于后续的查询
                part_code_column[i] = str(part_code_column[i])

            part_code = part_code_column[i]
            print(i, "  ", part_code)

            try:
                query_res = query_part_info(part_code, self.customer_code)

                # 如果有多个结果,记录到multi_results_map中
                if len(query_res) > 1:
                    multi_results_map[i] = query_res
                    print(f"料号 {part_code} 有 {len(query_res)} 条查询结果，将在后续处理中展开")

                # 先写入第一条结果
                self.query_res_process(part_code, query_res, i)
                print(f"完成查询: {part_code}")
            except Exception as e:
                print(f"查询料号 {part_code} 时出错: {e}")
                import traceback
                traceback.print_exc()
                # 出错时也要写入料号,避免卡住
                workbook = openpyxl.load_workbook('target.xlsx')
                sheet = workbook.active
                start_row_excel = 5
                sheet['D' + str(start_row_excel + i)] = part_code
                workbook.save('target.xlsx')

        # 如果有多条结果的料号，需要展开处理
        if multi_results_map:
            print(f"开始处理多条结果的料号，共 {len(multi_results_map)} 个")
            self.multi_results_process(multi_results_map)

    def query_res_process(self, part_code, query_res, part_index):
        """
        :param part_index:
        :param query_res: 查询结果列表
        :return:
        """
        # 处理查询结果列表，写入第一条结果
        if len(query_res) != 0:
            # 取第一条结果
            first_result = query_res[0]
            print(first_result['k3code'], first_result['type_name'], first_result['specification'])
            workbook = openpyxl.load_workbook('target.xlsx')
            sheet = workbook.active
            start_row = 5  # 开始写的行
            sheet['B' + str(start_row + part_index)] = first_result['k3code']

            sheet['C' + str(start_row + part_index)] = first_result['type_name']
            #现在改成不删除客户编码了，因为客户一致的时候本来就不删除 然后客户不一致的时候 也没必要删除 代表 需要工程手工加入k3库中
            sheet['D' + str(start_row + part_index)] = first_result['specification']

            # 写入匹配类型标记到O列(第15列)
            if 'match_type' in first_result:
                sheet['O' + str(start_row + part_index)] = first_result['match_type']

            workbook.save('target.xlsx')

        else:  # 无查询结果，在规格栏写上料号
            workbook = openpyxl.load_workbook('target.xlsx')
            sheet = workbook.active
            start_row = 5  # 开始写的行
            sheet['D' + str(start_row + part_index)] = part_code
            workbook.save('target.xlsx')

    def multi_results_process(self, multi_results_map):
        """
        处理有多条查询结果的料号，展开成多行
        使用openpyxl直接操作Excel，保持原有格式
        :param multi_results_map: {part_index: [result1, result2, ...]}
        """
        workbook = openpyxl.load_workbook('target.xlsx')
        sheet = workbook.active

        # 按照part_index从大到小排序，这样插入时不会影响前面的索引
        sorted_indices = sorted(multi_results_map.keys(), reverse=True)

        for part_index in sorted_indices:
            results_list = multi_results_map[part_index]
            print(f"处理料号索引 {part_index}，有 {len(results_list)} 条结果")

            # Excel中的实际行号(+5是因为前4行是表头,+1是因为Excel行号从1开始)
            excel_row = part_index + 5

            # 第一条结果不标记为multi,只有第2条及以后的才标记
            # 从第2条结果开始,在当前行下方插入新行
            for result_idx in range(1, len(results_list)):
                result = results_list[result_idx]

                # 在当前行下方插入一行
                insert_row = excel_row + result_idx
                sheet.insert_rows(insert_row)

                # 复制原始行的所有单元格到新行(包括格式)
                for col in range(1, sheet.max_column + 1):
                    source_cell = sheet.cell(row=excel_row, column=col)
                    target_cell = sheet.cell(row=insert_row, column=col)

                    # 复制值
                    target_cell.value = source_cell.value

                    # 复制格式
                    if source_cell.has_style:
                        target_cell.font = source_cell.font.copy()
                        target_cell.border = source_cell.border.copy()
                        target_cell.fill = source_cell.fill.copy()
                        target_cell.number_format = source_cell.number_format
                        target_cell.protection = source_cell.protection.copy()
                        target_cell.alignment = source_cell.alignment.copy()

                # 填写新行的K3信息
                sheet.cell(row=insert_row, column=2, value=result['k3code'])  # B列: K3 code
                sheet.cell(row=insert_row, column=3, value=result['type_name'])  # C列: 类型
                sheet.cell(row=insert_row, column=4, value=result['specification'])  # D列: 规格
                sheet.cell(row=insert_row, column=15, value='multi')  # O列: 多结果标记

                # 如果有匹配类型标记,也写入(但不覆盖multi标记)
                if 'match_type' in result:
                    sheet.cell(row=insert_row, column=15, value=f"multi-{result['match_type']}")  # O列: 多结果+匹配类型

                print(f"  添加第 {result_idx + 1} 条结果到Excel第 {insert_row} 行")

        # 保存Excel文件
        workbook.save('target.xlsx')
        print("多结果处理完成，已保存到target.xlsx")

    def DB_process(self):
        """
        读取用户输入的K3文件，建立查询的数据库
        :return:
        """
        K3Data.truncate_table()
        print(K3Data.select().count())
        excel_file_path = self.K3Input_textEdit.input_path_file
        df = pd.read_excel(excel_file_path)

        data = [{'k3code': row['K3Code'], 'type_name': row['Name'], 'specification': row['Specification']} for _, row in
                df.iterrows()]
        K3Data.bulk_create([K3Data(**row) for row in data], batch_size=100)  # 每批次插入100条数据
        # for index, row in df.iterrows():
        #   note = K3Data.create(k3code=row['K3Code'], type_name=row['Name'], specification=row['Specification'])
        # 将读入的K3BOM建立查询的数据库
        # K3Data.insert(df).execute()
        print(K3Data.select().count())

    def alternative_process(self):
        excel_file_path = self.BOMInput_textEdit.input_path_file
        df_initial_bom_path = pd.read_excel(excel_file_path)  # 读取原始BOM
        df_res_bom = pd.read_excel('target.xlsx', skiprows=4, header=None)  # 读取工程BOM

        # 将可能包含字符串的列转换为 object 类型，避免类型不兼容警告
        df_res_bom[1] = df_res_bom[1].astype('object')  # K3 code 列
        df_res_bom[2] = df_res_bom[2].astype('object')  # 类型列
        df_res_bom[3] = df_res_bom[3].astype('object')  # 规格列

        # 从原始BOM读取替代信息
        start_row = self.initial_row_line - 2  # 指定起始行索引（从0开始）
        end_row = self.end_row_line - 1
        alter1_code_list = df_initial_bom_path.iloc[start_row:end_row, self.alternative1_code_col].tolist()  # 替代1料号
        manu1_list = df_initial_bom_path.iloc[start_row:end_row, self.alternative1_manufacturer_col].tolist()  # 替代1厂商
        alter2_code_list = df_initial_bom_path.iloc[start_row:end_row, self.alternative2_code_col].tolist()  # 替代2料号
        manu2_list = df_initial_bom_path.iloc[start_row:end_row, self.alternative2_manufacturer_col].tolist()  # 替代2厂商

        # 建立原始BOM行号到工程BOM行号的正确映射
        # 需要跳过标记为'multi'的行,因为这些行是multi_results_process插入的
        res_bom_index = 0  # 工程BOM的当前行索引
        original_bom_row_count = 0  # 已经处理的原始BOM行数

        # 遍历原始BOM的每一行，处理替代料
        for original_index, alter1_code in enumerate(alter1_code_list):
            # 跳过所有标记为'multi'的行,找到当前原始BOM行对应的工程BOM行
            while res_bom_index < len(df_res_bom) and df_res_bom.iloc[res_bom_index, 14] == 'multi':
                res_bom_index += 1

            # 确保索引没有越界
            if res_bom_index >= len(df_res_bom):
                print(f"警告: res_bom_index={res_bom_index} 超出范围,工程BOM总行数={len(df_res_bom)}")
                break
            if not pd.isnull(alter1_code):
                print(f"处理替代1: original_index={original_index}, code={alter1_code}, res_bom_index={res_bom_index}")

                # 获取当前要复制的主料行
                row_to_copy = df_res_bom.iloc[res_bom_index:res_bom_index + 1].copy()  # 使用copy()避免警告

                # 插入替代1行
                df_res_bom = pd.concat([
                    df_res_bom.iloc[:res_bom_index + 1],
                    row_to_copy,
                    df_res_bom.iloc[res_bom_index + 1:]
                ], ignore_index=True)

                # 处理替代1
                res_bom_index += 1  # 移动到新插入的替代1行

                # 设置替代1厂商
                df_res_bom.iloc[res_bom_index, 8] = manu1_list[original_index]
                df_res_bom.iloc[res_bom_index, 14] = "替代1"
                df_res_bom.iloc[res_bom_index, 1] = None  # 清空K3 code

                print(f"开始查询替代1料号: {alter1_code}")
                query_res_alter1 = query_part_info(str(alter1_code), self.customer_code)
                print(f"替代1查询完成: {len(query_res_alter1)} 条结果")

                if len(query_res_alter1) != 0:
                    # 取第一条结果
                    first_res = query_res_alter1[0]
                    df_res_bom.iloc[res_bom_index, 1] = first_res['k3code']  # 替代1 k3code填写
                    df_res_bom.iloc[res_bom_index, 2] = first_res['type_name']  # 替代1 类型填写
                    df_res_bom.iloc[res_bom_index, 3] = first_res['specification'].replace(
                        self.customer_code + " ", "")  # 替代1 规格填写

                    # 检查主料有无查询结果，如果无，那么用替代1 信息复制过去
                    if pd.isnull(df_res_bom.iloc[res_bom_index - 1, 2]):
                        df_res_bom.iloc[res_bom_index - 1, 2] = first_res['type_name']
                        temp_part_code = df_res_bom.iloc[res_bom_index - 1, 3]  # 主料料号
                        temp_manu = df_res_bom.iloc[res_bom_index - 1, 8]  # 主料供应商
                        spec_parts = re.split(r'[ 　]+', first_res['specification'])
                        last_word_is_footprint = any_footprint_in_string(spec_parts[-1])  # 规格的最后一个单词 是否是封装

                        if last_word_is_footprint:  # 如果最后一个词是封装，直接复制过去即可
                            df_res_bom.iloc[res_bom_index - 1, 3] = first_res['specification'].replace(
                                self.customer_code + " ", "").replace(str(alter1_code), temp_part_code)+' '+temp_manu
                        else:  # 如果最后一个单词是品牌
                            df_res_bom.iloc[res_bom_index - 1, 3] = remove_last_word(
                                first_res['specification']).replace(
                                str(self.customer_code) + " ", "").replace(str(alter1_code), temp_part_code) + ' ' + str(temp_manu)

                else:  # 替代一查询无果，查询主料号记录，复制替代1里
                    if not pd.isnull(df_res_bom.iloc[res_bom_index - 1, 2]):
                        df_res_bom.iloc[res_bom_index, 2] = df_res_bom.iloc[res_bom_index - 1, 2]  # 使用主料号的类型
                        main_parts = re.split(r'[ 　]+', df_res_bom.iloc[res_bom_index - 1, 3])
                        temp_part_code = main_parts[0]  # 主料的料号
                        temp_manu = df_res_bom.iloc[res_bom_index, 8]  # 替代1 的品牌
                        last_word_is_footprint = any_footprint_in_string(main_parts[-1])  # 规格的最后一个单词 是否是封装
                        if last_word_is_footprint:
                            df_res_bom.iloc[res_bom_index, 3] = df_res_bom.iloc[res_bom_index - 1, 3].replace(
                                temp_part_code, str(alter1_code))+' '+temp_manu  # 替代1 规格填写
                        else:
                            df_res_bom.iloc[res_bom_index, 3] = remove_last_word(df_res_bom.iloc[res_bom_index - 1, 3]).replace(
                                temp_part_code, str(alter1_code))+' '+temp_manu  # 替代1 规格填写
                    else:
                        # 主料号也无查询结果，暂时填一个替代一的物料号
                        df_res_bom.iloc[res_bom_index, 3] = str(alter1_code)

                # 检查是否有替代2
                if self.alternative2_code_col != -1 and not pd.isnull(alter2_code_list[original_index]):
                    print(f"处理替代2: code={alter2_code_list[original_index]}")

                    # 再次复制主料行作为替代2的基础
                    row_to_copy2 = df_res_bom.iloc[res_bom_index - 1:res_bom_index].copy()  # 复制主料行
                    df_res_bom = pd.concat([
                        df_res_bom.iloc[:res_bom_index + 1],
                        row_to_copy2,
                        df_res_bom.iloc[res_bom_index + 1:]
                    ], ignore_index=True)

                    res_bom_index += 1  # 移动到新插入的替代2行

                    # 设置替代2厂商
                    df_res_bom.iloc[res_bom_index, 8] = manu2_list[original_index]
                    df_res_bom.iloc[res_bom_index, 14] = "替代2"
                    df_res_bom.iloc[res_bom_index, 1] = None  # 清空K3 code

                    query_res_alter2 = query_part_info(str(alter2_code_list[original_index]), self.customer_code)
                    if len(query_res_alter2) != 0:
                        # 取第一条结果
                        first_res2 = query_res_alter2[0]
                        df_res_bom.iloc[res_bom_index, 1] = first_res2['k3code']  # 替代2 k3code填写
                        df_res_bom.iloc[res_bom_index, 2] = first_res2['type_name']  # 替代2 类型填写
                        df_res_bom.iloc[res_bom_index, 3] = first_res2['specification'].replace(
                            self.customer_code + " ", "")  # 替代2 规格填写

                        # 检查替代1有无查询结果，如果无，那么用替代2填补
                        if pd.isnull(df_res_bom.iloc[res_bom_index - 1, 2]):
                            df_res_bom.iloc[res_bom_index - 1, 2] = first_res2['type_name']
                            temp_part_code = df_res_bom.iloc[res_bom_index - 1, 3]
                            temp_manu = df_res_bom.iloc[res_bom_index - 1, 8]
                            spec2_parts = re.split(r'[ 　]+', first_res2['specification'])
                            last_word_is_footprint = any_footprint_in_string(spec2_parts[-1])
                            if last_word_is_footprint:
                                df_res_bom.iloc[res_bom_index - 1, 3] = first_res2['specification'].replace(
                                    self.customer_code + " ", "").replace(str(alter2_code_list[original_index]), temp_part_code) +' '+temp_manu
                            else:
                                df_res_bom.iloc[res_bom_index - 1, 3] = remove_last_word(first_res2['specification']).replace(
                                    self.customer_code + " ", "").replace(str(alter2_code_list[original_index]), temp_part_code) +' '+temp_manu

                        # 检查主料有无查询结果，如果无，那么用替代2填补
                        if pd.isnull(df_res_bom.iloc[res_bom_index - 2, 2]):
                            df_res_bom.iloc[res_bom_index - 2, 2] = first_res2['type_name']
                            temp_part_code = df_res_bom.iloc[res_bom_index - 2, 3]
                            temp_manu = df_res_bom.iloc[res_bom_index - 2, 8]
                            spec2_parts2 = re.split(r'[ 　]+', first_res2['specification'])
                            last_word_is_footprint = any_footprint_in_string(spec2_parts2[-1])
                            if last_word_is_footprint:
                                df_res_bom.iloc[res_bom_index - 2, 3] = first_res2['specification'].replace(
                                    self.customer_code + " ", "").replace(str(alter2_code_list[original_index]), temp_part_code)+' '+temp_manu
                            else:
                                df_res_bom.iloc[res_bom_index - 2, 3] = remove_last_word(first_res2['specification']).replace(
                                    self.customer_code + " ", "").replace(str(alter2_code_list[original_index]), temp_part_code)+' '+temp_manu
                    else:
                        if not pd.isnull(df_res_bom.iloc[res_bom_index - 1, 2]):
                            df_res_bom.iloc[res_bom_index, 2] = df_res_bom.iloc[res_bom_index - 1, 2]  # 替代2 类型填写
                            main_parts2 = re.split(r'[ 　]+', df_res_bom.iloc[res_bom_index - 1, 3])
                            temp_part_code = main_parts2[0]
                            temp_manu = df_res_bom.iloc[res_bom_index, 8]
                            last_word_is_footprint = any_footprint_in_string(main_parts2[-1])
                            if last_word_is_footprint:
                                df_res_bom.iloc[res_bom_index, 3] = df_res_bom.iloc[res_bom_index - 1, 3].replace(
                                    temp_part_code, str(alter2_code_list[original_index]))+' '+ temp_manu  # 替代2 规格填写
                            else:
                                df_res_bom.iloc[res_bom_index, 3] = remove_last_word(df_res_bom.iloc[res_bom_index - 1, 3]).replace(
                                    temp_part_code, str(alter2_code_list[original_index]))+' '+ temp_manu  # 替代2 规格填写
                        else:
                            # 主料号也无查询结果，暂时填一个替代二的物料号
                            df_res_bom.iloc[res_bom_index, 3] = str(alter2_code_list[original_index])

                # 移动到下一个主料位置
                res_bom_index += 1
            else:
                # 没有替代料，直接移动到下一个主料
                res_bom_index += 1

            # 每处理完一个原始BOM行,记录已处理数量
            original_bom_row_count += 1

        # 使用openpyxl直接写入,避免创建新sheet
        workbook = openpyxl.load_workbook('output.xlsx')
        sheet = workbook.active

        # 从第5行开始写入数据
        start_row = 5
        for row_idx, row_data in df_res_bom.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=start_row + row_idx, column=col_idx, value=value)

        workbook.save('output.xlsx')
        print("替代料号处理完成,已保存到output.xlsx")

    def online_query(self):
        token = token_access()
        if self.alternative1_code_col == -1:
            query_online('target.xlsx', token, self.BOMInput_textEdit.input_path_file)
        else:
            query_online('output.xlsx', token, self.BOMInput_textEdit.input_path_file)
        print("process complete")


if __name__ == "__main__":
    app = QApplication([])
    window = MyMainWindow()
    window.show()
    app.exec()
