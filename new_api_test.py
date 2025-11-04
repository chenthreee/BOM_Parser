import hashlib
import json
import os
import time
import token
import urllib

import openpyxl
import pandas as pd
import requests
from openpyxl.styles import Font,PatternFill, Border
from openpyxl.workbook import Workbook

from query import token_access, url_encode, md5, query_res_process, IC_list, refresh_excel, color_mark
def single_query(url, token, part_code):
    temp = str(part_code)
    current_time = int(time.time())
    readytoMD5 = f"_t={current_time}"
    temp = temp.replace(" ", "/")  # Remove spaces
    readytoMD5 += f"&keyword={temp}"
    readytoMD5 += f"&token={token}"
    readytoMD5 += "|7b40ee9ed472baa1bc1a02a74827217a"

    readytoMD5 = url_encode(readytoMD5)
    sign = md5(readytoMD5)
    params = {
        '_t': current_time,
        'keyword': temp,
        'sign': sign,
        'token': token
    }
    # print(params)
    headers = {
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'
    }

    response = requests.post(url, data=params, headers=headers)
    if response.status_code != 200:
        return None

    try:
        # 解析 JSON 响应
        result = response.json().get("result", [])
    except requests.exceptions.JSONDecodeError as e:
        result = None
    # return response.json().get("result", [])

    print(response.json())
    return result


def quert_test(url, token, input_temp_file):
    headers = {
        "Content-Type": "application/json",
    }
    # 定义请求数据
    data = {
        "token": str(token),
        "_t": int(time.time()),
        "sign": "",
        "items": []
    }

    df_temp_bom = pd.read_excel(input_temp_file, skiprows=4, header=None)
    for index, row in df_temp_bom.iterrows():
        if pd.notna(row[2]):
            pass
        else:
            temp_part_num = row[3]
            temp_purchase_num = str(row[6])
            temp_part = {"keyword": temp_part_num, "pro_num": temp_purchase_num}
            data['items'].append(temp_part)

    # 将请求数据进行URL编码
    encoded_items = data['items']
    encoded_token = data['token']
    encoded_t = str(data['_t'])
    temp_str = ""
    index = 0

    for item in encoded_items:
        temp_str += "&items[" + str(index) + "][keyword]=" + str(item["keyword"])
        temp_str += "&items[" + str(index) + "][pro_num]=" + str(item["pro_num"])
        index = index + 1
    ready_to_md5 = f"_t={encoded_t}{temp_str}&token={encoded_token}|7b40ee9ed472baa1bc1a02a74827217a"
    ready_to_md5 = url_encode(ready_to_md5)
    sign = md5(ready_to_md5)
    data['sign'] = sign
    # 发送POST请求
    response = requests.post(url, headers=headers, json=data)
    return response


def quotation_query_test(url, token, row_code_list):
    # headers = {
    #     "Content-Type": "application/json",
    # }
    # 定义请求数据
    data = {
        "token": str(token),
        "_t": int(time.time()),
        "sign": "",
        "row_code": row_code_list
    }
    # 填充row_code

    encoded_token = data['token']
    encoded_t = str(data['_t'])
    temp_str = ""
    index = 0

    for row_code in row_code_list:
        temp_str += "&row_code[" + str(index) + "]=" + str(row_code)
        index = index + 1
    ready_to_md5 = f"_t={encoded_t}{temp_str}&token={encoded_token}|7b40ee9ed472baa1bc1a02a74827217a"
    ready_to_md5 = url_encode(ready_to_md5)
    sign = md5(ready_to_md5)
    data['sign'] = sign

    params = "?_t=" + str(data["_t"])
    params += "&token=" + str(token)
    params += "&sign=" + sign
    index = 0
    for row_code in row_code_list:
        params += "&row_code[" + str(index) + "]=" + str(row_code)
        index = index + 1

    # 发送Get请求
    full_url = url + params
    response = requests.get(full_url)
    return response


def row_code_process(temp_file_path, row_code_query_res):
    df_temp_bom = pd.read_excel(temp_file_path, skiprows=4, header=None)
    result_dict = {}
    i = 0
    for index, row in df_temp_bom.iterrows():
        if pd.notna(row[2]):
            pass
        else:
            temp_no = row[0]
            temp_row_code = row_code_query_res[i]['row_code']
            i = i + 1
            result_dict[temp_row_code] = temp_no

    return result_dict


def split_dict_keys_into_lists(d):
    keys = list(d.keys())  # 获取字典的所有键
    result = []
    temp_list = []

    for key in keys:
        temp_list.append(key)
        if len(temp_list) == 3:
            result.append(temp_list)
            temp_list = []

    # 将剩余的不足三个元素的列表添加到结果列表中
    if temp_list:
        result.append(temp_list)

    return result
def query_online(query_file, token, bom_file_path):
    df_temp_bom = pd.read_excel(query_file, skiprows=4, header=None)
    single_query_url = "https://openapi.ickey.cn/search-v1/products/get-single-goods-new"
    query_res = ''

    # cheng
    for index, row in df_temp_bom.iterrows():
        if pd.notna(row[2]):
            pass
        else:
            # 需要更换成新的接口

            query_res = single_query(single_query_url, token, row[3])

            if query_res is None:
                # On pending 如果查询不到，那么需要解析原始的
                print("Error: Failed to query.")
                # exit(1)
            else:
                # 对查询结果进行解析
                parser_res = query_res_process(query_res, row[3], row[8])

                # 向当前行的第一列写入parser_res中的字段
                if '贴片' in parser_res['part_type']: parser_res['mount_type'] = ''
                if '陶瓷' in parser_res['part_type']: parser_res['part_type'] = parser_res['part_type'].replace('陶瓷',
                                                                                                                '')
                for item in IC_list:
                    if item in parser_res['part_type']:
                        parser_res['part_type'] = 'IC'
                        break


                # 元件类型填入
                df_temp_bom.at[index, df_temp_bom.columns[2]] = parser_res['mount_type'] + parser_res['part_type']
                # 元件规格填入
                df_temp_bom.at[index, df_temp_bom.columns[3]] = str(row[3]) + ' ' + parser_res['specification'] + ' ' + \
                                                                parser_res['footprint']

                # 查询得到的描述粘贴
                df_temp_bom.at[index, df_temp_bom.columns[12]] = parser_res['pro_desc']
                # 查询得到的规格书链接
                if parser_res['datasheet_link']:
                    df_temp_bom.at[index, df_temp_bom.columns[13]] = 'https:' + parser_res['datasheet_link'][0]

            time.sleep(2)  # 每次查询完成 休眠两秒

    output_file = bom_file_path
    prefix = '自动生成工程BOM'
    name_part, extension = output_file.rsplit('.', 1)
    if extension == 'xls':
        extension = extension.replace('xls', 'xlsx')
    new_filename = name_part + prefix + '.' + extension

    if not os.path.exists(new_filename):
        source_filename = 'source.xlsx'
        source_wb = openpyxl.load_workbook(source_filename)
        source_ws = source_wb.active

        wb = Workbook()
        ws = wb.active

        # 复制前四行数据和样式
        for row in source_ws.iter_rows(min_row=1, max_row=4, max_col=source_ws.max_column):
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = Font(name=cell.font.name,
                                         size=cell.font.size,
                                         bold=cell.font.bold,
                                         italic=cell.font.italic,
                                         vertAlign=cell.font.vertAlign,
                                         underline=cell.font.underline,
                                         strike=cell.font.strike,
                                         color=cell.font.color)
                    new_cell.border = Border(left=cell.border.left,
                                             right=cell.border.right,
                                             top=cell.border.top,
                                             bottom=cell.border.bottom,
                                             diagonal=cell.border.diagonal,
                                             diagonal_direction=cell.border.diagonal_direction,
                                             outline=cell.border.outline,
                                             vertical=cell.border.vertical,
                                             horizontal=cell.border.horizontal)
                    new_cell.fill = PatternFill(fill_type=cell.fill.fill_type,
                                                start_color=cell.fill.start_color,
                                                end_color=cell.fill.end_color)
                    new_cell.number_format = cell.number_format
        wb.save(new_filename)

    refresh_excel(new_filename)
    with pd.ExcelWriter(new_filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df_temp_bom.to_excel(writer, index=False, startrow=4, header=False, sheet_name='Sheet')

    color_mark(new_filename)
    # df_temp_bom.to_excel(query_file, startrow=4, index=False, header=False)

temp_file_path = 'target.xlsx'
# url = "https://openapi.ickey.cn/search-v1/products/offer"
token_get = token_access()
# response = quert_test(url, token_get, temp_file_path)
# print(response.json())
#
# # 处理返回结果，存储在list中，序号保证到之后能写回去，
# res_collect = row_code_process(temp_file_path, response.json().get("result", []))
# dict_to_list = split_dict_keys_into_lists(res_collect)
#
# for dict_item in dict_to_list:
#     url = "https://openapi.ickey.cn/search-v1/products/polling-offer"
#     res = quotation_query_test(url, token_get, dict_item)
#     print(res.json())
query_online(temp_file_path, token_get,'test.xlsx')